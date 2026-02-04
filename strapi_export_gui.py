import json
import logging
import os
import sys
import threading
import queue
from datetime import datetime
from typing import Any, Dict, List, Tuple

import requests
import tkinter as tk
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tkinter import ttk, filedialog, messagebox


def _get_log_path() -> str:
    if getattr(sys, "frozen", False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.abspath(os.getcwd())
    logs_dir = os.path.join(base_dir, "logs")
    os.makedirs(logs_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(logs_dir, f"app_{timestamp}.log")


def _rotate_logs(logs_dir: str, max_files: int = 7) -> None:
    try:
        entries = [
            os.path.join(logs_dir, f)
            for f in os.listdir(logs_dir)
            if f.startswith("app_") and f.endswith(".log")
        ]
        entries.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        for old_path in entries[max_files:]:
            os.remove(old_path)
    except Exception:
        pass


logger = logging.getLogger(__name__)
if not logger.handlers:
    log_path = _get_log_path()
    logging.basicConfig(
        level=logging.INFO,
        filename=log_path,
        filemode="a",
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    _rotate_logs(os.path.dirname(log_path), max_files=7)


def _log_unhandled_exception(exc_type, exc, tb):
    logging.getLogger(__name__).error(
        "Unhandled exception", exc_info=(exc_type, exc, tb)
    )


sys.excepthook = _log_unhandled_exception


def _safe_json_dumps(value: Any) -> str:
    try:
        return json.dumps(value, ensure_ascii=False)
    except Exception:
        return str(value)


def _normalize_item(item: Dict[str, Any], flatten_attributes: bool) -> Dict[str, Any]:
    # Strapi v4 returns { id, attributes: {...} }
    # Strapi v5 can return flattened fields directly.
    if flatten_attributes and isinstance(item, dict) and "attributes" in item:
        attrs = item.get("attributes") or {}
        if not isinstance(attrs, dict):
            attrs = {"attributes": attrs}
        merged = {"id": item.get("id")}
        merged.update(attrs)
        return merged
    return item


def _relation_display(value: Dict[str, Any]) -> str:
    for key in ("name", "title", "label", "slug"):
        if key in value and value[key] not in (None, ""):
            return str(value[key])
    return _safe_json_dumps(value)


def _prepare_item_for_export(item: Dict[str, Any]) -> Dict[str, Any]:
    prepared: Dict[str, Any] = {}
    for key, value in item.items():
        if isinstance(value, dict):
            prepared[key] = _relation_display(value)
        elif isinstance(value, list):
            if not value:
                prepared[key] = ""
            elif all(isinstance(v, dict) for v in value):
                prepared[key] = ", ".join(_relation_display(v) for v in value)
            else:
                prepared[key] = ", ".join(str(v) for v in value)
        else:
            prepared[key] = value
    return prepared


def _extract_size_label(size_item: Dict[str, Any]) -> str:
    for key in ("EU", "size", "rozmiar", "label"):
        val = size_item.get(key)
        if val not in (None, ""):
            return str(val)
    return ""


def _expand_sizes(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    expanded: List[Dict[str, Any]] = []
    for item in items:
        sizes = item.get("rozmiary")
        if isinstance(sizes, list) and sizes:
            for size_item in sizes:
                new_item = dict(item)
                new_item.pop("rozmiary", None)
                if isinstance(size_item, dict):
                    size_label = _extract_size_label(size_item)
                    new_item["rozmiar"] = size_label
                    for key in ("EU", "US", "UK", "CM", "J", "SKU", "EAN", "Stock"):
                        if key in size_item:
                            new_item[f"rozmiar_{key}"] = size_item.get(key)
                else:
                    new_item["rozmiar"] = str(size_item)
                sku_val = new_item.get("sku")
                if sku_val and new_item.get("rozmiar"):
                    new_item["sku"] = f"{sku_val}/{new_item['rozmiar']}"
                expanded.append(new_item)
        else:
            new_item = dict(item)
            new_item.pop("rozmiary", None)
            new_item["rozmiar"] = ""
            expanded.append(new_item)
    return expanded


class StrapiExporterGUI:
    def __init__(self, root: tk.Tk) -> None:
        load_dotenv()
        self.root = root
        self.root.title("Strapi Exporter")
        self.root.geometry("760x520")

        self._queue: "queue.Queue[Tuple[str, Any]]" = queue.Queue()
        self._stop_flag = threading.Event()

        self._build_ui()
        self._load_settings()
        self._poll_queue()

    def _build_ui(self) -> None:
        main = ttk.Frame(self.root, padding="10")
        main.pack(fill=tk.BOTH, expand=True)

        config = ttk.LabelFrame(main, text="Konfiguracja Strapi", padding="10")
        config.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(config, text="URL Strapi:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.url_var = tk.StringVar(value=os.getenv("STRAPI_URL", "https://pim.flmsc.com.pl"))
        ttk.Entry(config, textvariable=self.url_var, width=55).grid(row=0, column=1, sticky=tk.W, padx=5, pady=2)

        ttk.Label(config, text="API Token:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.token_var = tk.StringVar(value=os.getenv("STRAPI_API_TOKEN", ""))
        self.token_entry = ttk.Entry(config, textvariable=self.token_var, width=55, show="*")
        self.token_entry.grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)

        ttk.Label(config, text="Endpoint:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.endpoint_var = tk.StringVar(value="/api/products")
        ttk.Entry(config, textvariable=self.endpoint_var, width=55).grid(row=2, column=1, sticky=tk.W, padx=5, pady=2)

        ttk.Label(config, text="Populate:").grid(row=3, column=0, sticky=tk.W, pady=2)
        self.populate_var = tk.StringVar(value="*")
        ttk.Entry(config, textvariable=self.populate_var, width=55).grid(row=3, column=1, sticky=tk.W, padx=5, pady=2)

        ttk.Label(config, text="Limit strony:").grid(row=4, column=0, sticky=tk.W, pady=2)
        self.limit_var = tk.StringVar(value="100")
        ttk.Entry(config, textvariable=self.limit_var, width=10).grid(row=4, column=1, sticky=tk.W, padx=5, pady=2)

        options = ttk.LabelFrame(main, text="Opcje eksportu", padding="10")
        options.pack(fill=tk.X, pady=(0, 10))

        self.flatten_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options, text="Spłaszcz attributes (Strapi v4)", variable=self.flatten_var).pack(side=tk.LEFT, padx=10)

        actions = ttk.LabelFrame(main, text="Akcje", padding="10")
        actions.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(actions, text="Testuj połączenie", command=self._test_connection).pack(side=tk.LEFT, padx=5)
        ttk.Button(actions, text="Pobierz wszystkie produkty", command=self._start_export).pack(side=tk.LEFT, padx=5)
        ttk.Button(actions, text="Przerwij", command=self._stop_export).pack(side=tk.LEFT, padx=5)

        self.progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(main, variable=self.progress_var, maximum=100).pack(fill=tk.X, pady=(0, 10))

        self.status_var = tk.StringVar(value="Gotowy")
        ttk.Label(main, textvariable=self.status_var).pack(anchor=tk.W)

        log_frame = ttk.LabelFrame(main, text="Log", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = tk.Text(log_frame, height=12, state=tk.DISABLED)
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def _settings_path(self) -> str:
        return os.path.join(os.getcwd(), "settings.json")

    def _load_settings(self) -> None:
        path = self._settings_path()
        if not os.path.exists(path):
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if "url" in data:
                self.url_var.set(data["url"])
            if "token" in data:
                self.token_var.set(data["token"])
            if "endpoint" in data:
                self.endpoint_var.set(data["endpoint"])
            if "populate" in data:
                self.populate_var.set(data["populate"])
            if "limit" in data:
                self.limit_var.set(str(data["limit"]))
            if "flatten" in data:
                self.flatten_var.set(bool(data["flatten"]))
        except Exception:
            logger.exception("Failed to load settings")

    def _save_settings(self) -> None:
        data = {
            "url": self.url_var.get().strip(),
            "token": self.token_var.get(),
            "endpoint": self.endpoint_var.get().strip(),
            "populate": self.populate_var.get().strip(),
            "limit": self.limit_var.get().strip(),
            "flatten": bool(self.flatten_var.get()),
        }
        try:
            with open(self._settings_path(), "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception:
            logger.exception("Failed to save settings")

    def _log(self, message: str) -> None:
        self.log_text.config(state=tk.NORMAL)
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{ts}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update_idletasks()
        logger.info(message)

    def _headers(self) -> Dict[str, str]:
        token = self.token_var.get().strip()
        headers = {"Content-Type": "application/json"}
        if token:
            headers["Authorization"] = f"Bearer {token}"
        return headers

    def _build_url(self, start: int, limit: int) -> str:
        base = self.url_var.get().rstrip("/")
        endpoint = self.endpoint_var.get().lstrip("/")
        populate = self.populate_var.get().strip()
        query = f"?pagination[start]={start}&pagination[limit]={limit}"
        if populate:
            query += f"&populate={populate}"
        return f"{base}/{endpoint}{query}"

    def _test_connection(self) -> None:
        self._save_settings()
        try:
            url = self._build_url(0, 1)
            self._log(f"Testuję: {url}")
            resp = requests.get(url, headers=self._headers(), timeout=20)
            if resp.status_code == 200:
                total = resp.json().get("meta", {}).get("pagination", {}).get("total", "?")
                self._log(f"OK. Produkty: {total}")
                messagebox.showinfo("Sukces", f"Połączenie OK. Produkty: {total}")
            else:
                self._log(f"Błąd: {resp.status_code} - {resp.text[:200]}")
                messagebox.showerror("Błąd", f"HTTP {resp.status_code}")
        except Exception as exc:
            self._log(f"Błąd połączenia: {exc}")
            logger.exception("Connection error")
            messagebox.showerror("Błąd", str(exc))

    def _start_export(self) -> None:
        if self._stop_flag.is_set():
            self._stop_flag.clear()

        self._save_settings()
        default_name = f"strapi_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=default_name,
        )
        if not filename:
            return

        try:
            limit = int(self.limit_var.get())
            if limit <= 0:
                raise ValueError()
        except Exception:
            messagebox.showerror("Błąd", "Limit strony musi być liczbą dodatnią.")
            return

        t = threading.Thread(target=self._export_worker, args=(filename, limit), daemon=True)
        t.start()

    def _stop_export(self) -> None:
        self._stop_flag.set()
        self._log("Zatrzymywanie...")

    def _export_worker(self, filename: str, limit: int) -> None:
        self._queue.put(("status", "Pobieranie..."))
        self._queue.put(("progress", 0))
        self._queue.put(("log", "Start pobierania produktów..."))

        all_items: List[Dict[str, Any]] = []
        start = 0
        total = None

        while True:
            if self._stop_flag.is_set():
                self._queue.put(("log", "Przerwano przez użytkownika."))
                self._queue.put(("status", "Przerwano"))
                return

            url = self._build_url(start, limit)
            try:
                resp = requests.get(url, headers=self._headers(), timeout=60)
            except Exception as exc:
                self._queue.put(("log", f"Błąd pobierania: {exc}"))
                logger.exception("Fetch error")
                self._queue.put(("status", "Błąd"))
                return

            if resp.status_code != 200:
                self._queue.put(("log", f"HTTP {resp.status_code}: {resp.text[:200]}"))
                self._queue.put(("status", "Błąd"))
                return

            data = resp.json()
            if total is None:
                total = data.get("meta", {}).get("pagination", {}).get("total", 0)
                self._queue.put(("log", f"Łącznie produktów: {total}"))

            items = data.get("data", [])
            if not items:
                break

            for item in items:
                all_items.append(_normalize_item(item, self.flatten_var.get()))

            start += limit
            if total:
                progress = min(100.0, (start / total) * 100)
                self._queue.put(("progress", progress))
                self._queue.put(("status", f"Pobrano {min(start, total)}/{total}"))
                self._queue.put(("log", f"Pobrano {min(start, total)}/{total}"))

            if total is not None and start >= total:
                break

        if not all_items:
            self._queue.put(("log", "Brak produktów do zapisu."))
            self._queue.put(("status", "Brak danych"))
            return

        all_items = _expand_sizes(all_items)
        all_items = [_prepare_item_for_export(item) for item in all_items]

        try:
            self._write_xlsx(filename, all_items)
        except Exception as exc:
            self._queue.put(("log", f"Błąd zapisu: {exc}"))
            logger.exception("Save error")
            self._queue.put(("status", "Błąd zapisu"))
            return

        self._queue.put(("progress", 100))
        self._queue.put(("status", "Gotowy"))
        self._queue.put(("log", f"Zapisano {len(all_items)} produktów do {filename}"))
        self._queue.put(("done", len(all_items)))

    def _write_xlsx(self, filename: str, items: List[Dict[str, Any]]) -> None:
        keys = []
        key_set = set()
        for item in items:
            for k in item.keys():
                if k not in key_set:
                    key_set.add(k)
                    keys.append(k)

        wb = Workbook()
        ws = wb.active
        ws.title = "Produkty"

        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=1, column=col_idx, value=key)

        for row_idx, item in enumerate(items, 2):
            for col_idx, key in enumerate(keys, 1):
                val = item.get(key, "")
                if isinstance(val, (dict, list)):
                    val = _safe_json_dumps(val)
                ws.cell(row=row_idx, column=col_idx, value=val)

        for col_idx, key in enumerate(keys, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(key))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell_val = row[0].value
                if cell_val is None:
                    continue
                max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        wb.save(filename)

    def _poll_queue(self) -> None:
        try:
            while True:
                msg_type, payload = self._queue.get_nowait()
                if msg_type == "log":
                    self._log(payload)
                elif msg_type == "status":
                    self.status_var.set(payload)
                elif msg_type == "progress":
                    self.progress_var.set(payload)
                elif msg_type == "done":
                    messagebox.showinfo("Zakończono", f"Wyeksportowano {payload} produktów.")
        except queue.Empty:
            pass
        self.root.after(100, self._poll_queue)


def main() -> None:
    root = tk.Tk()
    app = StrapiExporterGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
